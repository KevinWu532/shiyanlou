#!/usr/bin/env python3

from openpyxl import load_workbook, Workbook
from datetime import datetime

fb = load_workbook(filename='courses.xlsx')
sheet_students = fb['students']
sheet_time = fb['time']

def combine():
    sheet_combine = fb.create_sheet(title='combine')
    sheet_combine.append(['创建时间','课程名称','学习人数','学习时间'])
    for x in sheet_students.values:
        if x[2] != '学习人数':
            for time in sheet_time.values:
                if time[1] == x[1]:
                    sheet_combine.append(list(x)+[time[2]])
    fb.save('courses.xlsx')

def split():
    sheet_combine = fb['combine']
    time = []
    for x in sheet_combine.values:
        if x[0] != '创建时间':
            time.append(x[0].strftime('%Y'))

    for tm in set(time):
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title=tm)
        ws.append(['创建时间','课程名称','学习人数','学习时间'])
        for x in sheet_combine.values:
            if x[0] != '创建时间':
                if x[0].strftime('%Y') == tm:
                    ws.append(x)
        wb.save('{}.xlsx'.format(tm))

if __name__ == '__main__':
    combine()
    split()
